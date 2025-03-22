import sys
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def write_to_excel(output_excel_path, tables):
    wb = Workbook()

    # Удаляем дефолтную пустую страницу, чтобы потом создавать по одной на каждый table
    while len(wb.worksheets) > 0:
        wb.remove(wb.worksheets[0])

    for idx, table in enumerate(tables, start=1):
        sheet_name = f"Таблица_{idx}"
        ws = wb.create_sheet(title=sheet_name)

        # Получаем все строки данной таблицы
        rows = table.find_all("tr")
        # Матрица для хранения инфо о занятых ячейках
        occupied = {}

        current_row = 1
        for row in rows:
            # Получаем все столбцы (теги td и th) в строке
            cells = row.find_all(["td", "th"])
            current_col = 1
            for cell in cells:
                # Пропускаем уже занятые ячейки
                while (current_row, current_col) in occupied:
                    current_col += 1

                rowspan = int(cell.get("rowspan", 1))
                colspan = int(cell.get("colspan", 1))

                # Мёржим ячейки, если указан rowspan/colspan
                if rowspan > 1 or colspan > 1:
                    start_row = current_row
                    start_col = current_col
                    end_row = current_row + rowspan - 1
                    end_col = current_col + colspan - 1
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=start_col,
                        end_row=end_row,
                        end_column=end_col
                    )

                    # Помечаем все объединённые ячейки как занятые
                    for r in range(start_row, end_row + 1):
                        for c in range(start_col, end_col + 1):
                            occupied[(r, c)] = True

                value = cell.get_text(strip=True)
                ws.cell(row=current_row, column=current_col, value=value)

                current_col += 1

            current_row += 1

    wb.save(output_excel_path)

def get_tables(html_path, format_table):
    if format_table == "file":
        with open(html_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f, "html.parser")
    tables = soup.find_all("table")
    return tables

def check(tables):
    genuine_tables=[]
    for table in tables:
        if not(table.find("table")):
            genuine_tables+=[table]
    return genuine_tables
    
    
def data_acquisition():
    if len(sys.argv) < 3:
        print("Использование: python script.py url/file путь_к_html путь_к_xlsx")
        sys.exit(1)
    format_table = sys.argv[1]
    html_path = sys.argv[2]
    xlsx_path = sys.argv[3]
    
    if not os.path.isfile(html_path):
        print(f"Файл {html_path} не найден.")
        sys.exit(1)

    tables = get_tables(html_path, format_table) 
    
    return {'html_path': html_path, 'xlsx_path': xlsx_path, 'tables': tables}

def main():
    path = data_acquisition()
    print(path['tables'])
    genuine_tables = check(path['tables'])
    write_to_excel(path['xlsx_path'], genuine_tables)
    

if __name__ == "__main__":
    main()

    

    
