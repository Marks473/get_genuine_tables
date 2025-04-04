import sys
import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def write_to_excel(output_excel_path, tables):
    wb = Workbook()

    # Удаляем дефолтную пустую страницу, чтобы потом создавать по одной на каждый table
    while len(wb.worksheets) > 0:
        wb.remove(wb.worksheets[0])

    for idx, table in enumerate(tables, start=1):
        sheet_name = f"Таблица_{idx}"
        ws = wb.create_sheet(title=sheet_name)

        # Получаем все строки данной таблицы
        rows = table.find_all("tr")
        # Матрица для хранения инфо о занятых ячейках
        occupied = {}

        current_row = 1
        for row in rows:
            # Получаем все столбцы (теги td и th) в строке
            cells = row.find_all(["td", "th"])
            current_col = 1
            for cell in cells:
                # Пропускаем уже занятые ячейки
                while (current_row, current_col) in occupied:
                    current_col += 1

                rowspan = int(cell.get("rowspan", 1))
                colspan = int(cell.get("colspan", 1))

                # Мёржим ячейки, если указан rowspan/colspan
                if rowspan > 1 or colspan > 1:
                    start_row = current_row
                    start_col = current_col
                    end_row = current_row + rowspan - 1
                    end_col = current_col + colspan - 1
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=start_col,
                        end_row=end_row,
                        end_column=end_col
                    )

                    # Помечаем все объединённые ячейки как занятые
                    for r in range(start_row, end_row + 1):
                        for c in range(start_col, end_col + 1):
                            occupied[(r, c)] = True

                value = cell.get_text(strip=True)
                ws.cell(row=current_row, column=current_col, value=value)

                current_col += 1

            current_row += 1

    wb.save(output_excel_path)

def get_tables(html_path, format_table):
    if format_table == "file":

        if not os.path.isfile(html_path):
            print(f"Файл {html_path} не найден.")
            sys.exit(1)
        
        with open(html_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f, "html.parser")
        
    elif format_table == "url":
        f = download_html(html_path)
        soup = BeautifulSoup(f, "html.parser")
    else:
        print("воходные данные не коректны")
        sys.exit(1)
    tables = soup.find_all("table")
    return tables

def download_html(url):
    """Скачивание HTML содержимого по заданному URL."""
    response = requests.get(url)
    response.raise_for_status()  # Проверка успешности запроса
    return response.text

def get_tables_with_colspan_and_rowspan(table):
    
    rows_data = []

    # Итерируемся по всем строкам таблицы
    for row in table.find_all("tr"):
        row_cells = []
        # Ищем ячейки как <td>, так и <th>
        for cell in row.find_all(["td", "th"]):
            # Если атрибутов нет, считаем их равными 1
            rowspan = int(cell.get("rowspan", 1))
            colspan = int(cell.get("colspan", 1))
            row_cells.append({'cell': cell, 'rowspan': rowspan, 'colspan': colspan})
        rows_data.append(row_cells)
    return rows_data

def _chek(table_spans):
    if len(table_spans)<2:
        return False
    old = []
    new = []
    for cell in table_spans[0]:
        old+=[cell]
        if (cell["rowspan"] > 1):
            cell["rowspan"]-=1
            new+=[cell]
            
    for i in range(1,len(table_spans)):
        j=0
        k=0
        s=0
        while (j<len(new)) and (k < len(old)):
            print('new:',new)
            print('old:', old)
            if (new[j]["colspan"]) == old[k]["colspan"]:
                k += 1
                j += 1
            else:
                return False
        j=0
        print('прошло')
        while (j < len(table_spans[i])) and (k < len(old)):
            if (s + table_spans[i][j]["colspan"]) < old[k]["colspan"]:
                s += table_spans[i][j]["colspan"]
                new += [table_spans[i][j]]
                j += 1
            elif (s + table_spans[i][j]["colspan"]) == old[k]["colspan"]:
                #if not((old[k]["rowspan"] == 1) or (old[k]["rowspan"] == len(table_spans))):
                    #return False
                
                s = 0
                if (old[k]["rowspan"] > 1):
                    old[k]["rowspan"] -= 1
                    new += [old[k]]
                new += [table_spans[i][j]]
                j += 1
                k += 1
            else:
                return False
        if not((j >= len(table_spans[i])) and (k >= len(old))):
            return False
        old = [cell for cell in new]
        new =[]
        for cell in old:
            if (cell["rowspan"] > 1):
                cell["rowspan"] -= 1
                new += [cell]
    if len(new) != 0:
        return False
    return True
                

def chek(table):
    
    table_spans = get_tables_with_colspan_and_rowspan(table)
    
        
    
    return _chek(table_spans)

def get_genuine_tables(tables):
    genuine_tables=[]
    for table in tables:
        if chek(table) and not(table.find("table")):
            genuine_tables+=[table]
    return genuine_tables
    
    
def data_acquisition():
    if len(sys.argv) < -1:#3:
        print("Использование: python script.py url/file путь_к_html путь_к_xlsx")
        sys.exit(1)
    format_table = 'url' #sys.argv[1]
    html_path = 'https://2019.ecology-gosdoklad.ru/report/15/'#sys.argv[2]
    xlsx_path = 'example.xlsx'#sys.argv[3]

    tables = get_tables(html_path, format_table) 
    
    return {'html_path': html_path, 'xlsx_path': xlsx_path, 'tables': tables}
    
    
if __name__ == "__main__":
    path = data_acquisition()
    genuine_tables = get_genuine_tables(path['tables'])
    write_to_excel(path['xlsx_path'], genuine_tables)
    os.startfile(path['xlsx_path'])

    

    
