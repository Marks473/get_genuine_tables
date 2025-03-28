# -*- coding: utf-8 -*-
import sys
import os
from argparse import ArgumentError
import requests
from requests.packages import urllib3
from bs4 import BeautifulSoup
from openpyxl import Workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
class SpanTable:
    """таблица структуры row и col span и содержания таблицы"""

    def __init__(self):
        self.table = None

    def make_table(self, table):
        """Получение html таблицы и создания структуры"""

        table_span = []

        for row in table.find_all("tr"):

            if row.decode_contents() == '':
                continue

            row_cells = []
            for cell in row.find_all(["td", "th"]):
                rowspan = int(cell.get("rowspan", 1))
                colspan = int(cell.get("colspan", 1))
                row_cells += [{
                                'value': cell,
                                'rowspan': rowspan,
                                'colspan': colspan,
                                'rowspan_original': rowspan,
                                'colspan_original': colspan,
                                'similarity': False
                                }]

            table_span.append(row_cells)
        self.table = table_span

    def set_table(self, table):
        """Задать таблицу в той же структуре"""

        self.table = table

    def get_table(self):
        """Получение структуры таблицы"""

        return self.table

    def get_copy(self):
        """Получить копию структуры таблицы"""

        copy = []
        for i in range(len(self.table)):
            copy += [[]]
            for j in range(len(self.table[i])):
                cell = self.table[i][j]
                copy[i] += [{
                             'value': cell['value'],
                             'rowspan': cell['rowspan'],
                             'colspan': cell['colspan'],
                             'rowspan_original': cell['rowspan_original'],
                             'colspan_original': cell['colspan_original'],
                             'similarity': cell['similarity']
                             }]
        table_span = SpanTable()
        table_span.set_table(copy)
        return table_span

    def get_flip(self):
        """Получить таблицу отраженную относительно вертикали"""

        flip = self.get_copy().get_table()
        for i in range(len(flip)):
            flip[i] = flip[i][::-1]

        table_span = SpanTable()
        table_span.set_table(flip)
        return table_span

    def get_transpose(self):
        """Получить транспонированную таблицу"""

        rows_data = self.get_table()

        if not rows_data or not rows_data[0]:
            return []

        # Определяем размеры исходной таблицы
        rows = len(rows_data)
        cols = sum(cell['colspan'] for cell in rows_data[0])

        # Создаем матрицу для отслеживания занятых ячеек
        matrix = [[None for _ in range(cols)] for _ in range(rows)]

        # Заполняем матрицу данными и отмечаем занятые ячейки
        for i, row in enumerate(rows_data):
            current_col = 0
            for cell in row:
                # Находим следующую свободную позицию
                while current_col < cols and matrix[i][current_col] is not None:
                    current_col += 1

                # Заполняем ячейки согласно rowspan и colspan
                for r in range(cell['rowspan']):
                    for c in range(cell['colspan']):
                        if i + r < rows and current_col + c < cols:
                            matrix[i + r][current_col + c] = cell

                current_col += cell['colspan']

        # Транспонируем матрицу
        transposed_matrix = list(map(list, zip(*matrix)))

        # Создаем новую структуру данных
        transposed_data = []
        processed_cells = set()

        for i, row in enumerate(transposed_matrix):
            new_row = []
            for j, cell in enumerate(row):
                if cell is not None and (i, j) not in processed_cells:
                    # Создаем новую ячейку с поменянными местами rowspan и colspan
                    new_cell = {
                        'value': cell['value'],
                        'rowspan': cell['colspan'],  # Меняем местами
                        'colspan': cell['rowspan'],  # Меняем местами
                        'rowspan_original': cell['colspan_original'],  # Меняем местами
                        'colspan_original': cell['rowspan_original'],   # Меняем местами
                        'similarity': cell['similarity']
                    }
                    new_row.append(new_cell)

                    # Отмечаем обработанные ячейки
                    for r in range(new_cell['rowspan']):
                        for c in range(new_cell['colspan']):
                            processed_cells.add((i + r, j + c))

            if new_row:  # Добавляем строку только если в ней есть ячейки
                transposed_data.append(new_row)

        table_span = SpanTable()
        table_span.set_table(transposed_data)
        return table_span

    def get_type_of_genuine(self):
        if self.is_top():
            return 'top'
        if self.is_left():
            return 'left'
        if self.is_right():
            return 'right'
        if self.is_bottom():
            return 'bottom'
        return "not"

    def get_tag_structure(self, element):
        """Получает структуру тегов элемента в виде списка"""

        structure = []

        if isinstance(element, str):
            return None
        if element.name == 'span':
            return ''
        structure.append(element.name)

        for child in element.children:
            if isinstance(child, str) and child.strip() == '':
                continue

            child_structure = self.get_tag_structure(child)

            if child_structure:
                structure.append(child_structure)

        return structure

    def vertical_check(self):
        """Проверка на подлиность вертикальных таблиц"""

        table_spans = self.get_copy().get_table()
        if len(table_spans) < 2:
            return False

        old = [cell for cell in table_spans[0]]
        new = []

        for i in range(1, len(table_spans)):
            j = 0
            s = 0

            for k in range(len(old)):
                old[k]["rowspan"] -= 1

            k = 0

            while k < len(old):

                if old[k]["rowspan"] > 0:
                    new += [old[k]]
                    k += 1
                    continue

                if j < len(table_spans[i]):
                    for old_cell in old:
                        #print(old_cell["rowspan"], old_cell["rowspan_original"], table_spans[i][j]["rowspan"], '||||', old_cell["rowspan_original"] - old_cell["rowspan"], table_spans[i][j]["rowspan"] - old_cell["rowspan"] )
                        if (old_cell["rowspan"] > 0) and (old_cell["rowspan_original"] - old_cell["rowspan"] > 0 ) and (table_spans[i][j]["rowspan"] - old_cell["rowspan"]  > 0):
                            return False

                    if (s + table_spans[i][j]["colspan"]) < old[k]["colspan"]:
                        s += table_spans[i][j]["colspan"]
                        new += [table_spans[i][j]]
                        j += 1
                        continue

                    if (s + table_spans[i][j]["colspan"]) == old[k]["colspan"]:

                        if s == 0:
                            table_spans[i][j]['similarity'] = True
                            if old[k]["similarity"]:
                                old_structure = self.get_tag_structure(old[k]['value'])
                                new_structure = self.get_tag_structure(table_spans[i][j]['value'])
                                if not (old_structure == new_structure):
                                    return False
                        s = 0
                        new += [table_spans[i][j]]
                        j += 1
                        k += 1
                        continue

                return False

            if not ((j == len(table_spans[i])) and (k == len(old))):
                return False
            old = []
            old = [cell for cell in new]
            new = []

        if not (all(cell['rowspan'] == 1 for cell in old)):
            return False
        return True

    def is_top(self):
        return self.vertical_check()

    def is_left(self):
        table_span = self.get_transpose()
        return table_span.vertical_check()

    def is_right(self):
        table_span = self.get_flip().get_transpose()
        return table_span.vertical_check()

    def is_bottom(self):
        table_span = self.get_transpose().get_flip().get_transpose()
        return table_span.vertical_check()



def write_to_excel(output_excel_path, tables_span):
    """Запись таблицы в excel-фил"""

    if not(tables_span):
        return
    wb = Workbook()

    # Удаляем дефолтную пустую страницу, чтобы потом создавать по одной на каждый table
    while len(wb.worksheets) > 0:
        wb.remove(wb.worksheets[0])

    for idx, table_span in enumerate(tables_span, start=1):
        sheet_name = f"Таблица_{idx}"
        ws = wb.create_sheet(title=sheet_name)


        # Матрица для хранения инфо о занятых ячейках
        occupied = {}

        current_row = 1
        for row in table_span.get_table():
            if len(row) == 0:
                continue
            # Получаем все столбцы (теги td и th) в строке
            current_col = 1
            for cell in row:
                # Пропускаем уже занятые ячейки
                while (current_row, current_col) in occupied:
                    current_col += 1

                rowspan = cell['rowspan']
                colspan = cell['colspan']

                # Мёржим ячейки, если указан rowspan/colspan
                if rowspan > 1 or colspan > 1:
                    start_row = current_row
                    start_col = current_col
                    end_row = current_row + rowspan - 1
                    end_col = current_col + colspan - 1
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=start_col,
                        end_row=end_row,
                        end_column=end_col
                    )

                    # Помечаем все объединённые ячейки как занятые
                    for r in range(start_row, end_row + 1):
                        for c in range(start_col, end_col + 1):
                            occupied[(r, c)] = True

                value = cell['value'].get_text(strip=True)
                ws.cell(row=current_row, column=current_col, value=value)

                current_col += 1

            current_row += 1

    wb.save(output_excel_path)

def get_tables(html_path, format_table):
    """Извлечение таблицы"""
    if format_table == "file":

        if not os.path.isfile(html_path):
            print(f"Файл {html_path} не найден.")
            sys.exit(1)

        with open(html_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f, "html.parser")

    elif format_table == "url":
        f = download_html(html_path)
        soup = BeautifulSoup(f, "html.parser")
    else:
        print("входные данные не коректны")
        sys.exit(1)
    tables = soup.find_all("table", recursive=True)
    return tables

def download_html(url):
    """Скачивание HTML содержимого по-заданному URL."""
    response = requests.get(url, verify=False)
    response.raise_for_status()  # Проверка успешности запроса
    return response.content.decode("utf-8")

def get_genuine_tables(tables):
    """Получение все подленных таблиц"""
    genuine_tables = []
    for i, table in enumerate(tables):
        table_span = SpanTable()
        table_span.make_table(table)
        type_of_genuine = table_span.get_type_of_genuine()
        print(i + 1, type_of_genuine)

        if type_of_genuine != 'not':
            genuine_tables += [table_span]
    return genuine_tables

def arg_parser(args):
    if len(args) < 2:
        raise ArgumentError("Incorrect number of arguments")
    else:
        format_table = None
        html_path = None
        xlsx_path = None
        if args[1] == "-file":
            format_table = "file"
        elif args[1] == "-url":
            format_table = "url"
        else:
            raise ArgumentError("Invalid type of source")
        html_path = args[2]
        result = {"format_table": format_table, "html_path": html_path}
        if len(args) > 3:
            result.update({"xlsx_path": args[3]})
    return result

def data_acquisition():
    """Получение входных данных"""
    format_table = 'file'  # sys.argv[1]
    html_path = 'table.html'  # sys.argv[2]
    xlsx_path = 'example.xlsx'  # sys.argv[3]
    args_count = len(sys.argv)

    if args_count == 2 or args_count > 4:
        print("Использование: python script.py url/file путь_к_html путь_к_xlsx")
        raise ArgumentError("Incorrect arguments")
        # sys.exit(1)

    elif args_count > 1:
        # Аргументы есть
        a = arg_parser(sys.argv)
        format_table = a["format_table"]
        html_path = a["html_path"]
        if "xlsx_path" in a.keys():
            xlsx_path = a["xlsx_path"]

    tables = get_tables(html_path, format_table)

    return {'html_path': html_path, 'xlsx_path': xlsx_path, 'tables': tables}



if __name__ == "__main__":
    data = data_acquisition()
    genuine_tables = get_genuine_tables(data['tables'])
    write_to_excel(data['xlsx_path'], genuine_tables)
    os.startfile(data['xlsx_path'])
