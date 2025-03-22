import sys
import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter



class SpanTable:
    """таблица структуры row и col span и содержания таблицы"""
    
    def __init__(self):
        self.table = None

    def make_table(self, table):
        """Получение html таблицы и создания структуры"""

        table_span = []

        for row in table.find_all("tr"):

            if row.decode_contents() == '':
                continue
            
            row_cells = []
            for cell in row.find_all(["td", "th"]):
                rowspan = int(cell.get("rowspan", 1))
                colspan = int(cell.get("colspan", 1))
                row_cells.append({'value': cell, 'rowspan': rowspan, 'colspan': colspan, 'nottitle': False})

            table_span.append(row_cells)
        self.table = table_span

    def set_table(self, table):
        """Задать таблицу в той же структуре"""

        self.table = table

    def get_table(self):
        """Получение структуры таблицы"""

        return self.table

    def get_copy(self):
        """Получить копию структуры таблицы"""

        copy = []
        for i in range(len(self.table)):
            copy += [[]]
            for j in range(len(self.table[i])):
                cell = self.table[i][j]  # Исправлено: используем self.table вместо table_span
                value = cell['value']
                rowspan = cell['rowspan']
                colspan = cell['colspan']
                nottitle = cell['nottitle']
                copy[i] += [{'value': value,
                           'rowspan': rowspan,
                           'colspan': colspan,
                           'nottitle': nottitle}]  # Исправлено: используем nottitle из оригинальной ячейки
        table_span = SpanTable()
        table_span.set_table(copy)
        return table_span
    
    def get_flip(self):
        """Получить таблицу отраженную относительно вертикали"""

        flip = self.get_copy().table
        for i in range(len(self.table)):
            flip[i] = flip[i][::-1]
            
        table_span = SpanTable()
        table_span.set_table(flip)
        return table_span

    def get_transpose(self):
        """Получить транспонированную таблицу"""

        rows_data = self.get_table()
        
        if not rows_data or not rows_data[0]:
            return []

        # Определяем размеры исходной таблицы
        rows = len(rows_data)
        cols = sum(cell['colspan'] for cell in rows_data[0])

        # Создаем матрицу для отслеживания занятых ячеек
        matrix = [[None for _ in range(cols)] for _ in range(rows)]
        
        # Заполняем матрицу данными и отмечаем занятые ячейки
        for i, row in enumerate(rows_data):
            current_col = 0
            for cell in row:
                # Находим следующую свободную позицию
                while current_col < cols and matrix[i][current_col] is not None:
                    current_col += 1
                    
                # Заполняем ячейки согласно rowspan и colspan
                for r in range(cell['rowspan']):
                    for c in range(cell['colspan']):
                        if i + r < rows and current_col + c < cols:
                            matrix[i + r][current_col + c] = cell
                
                current_col += cell['colspan']

        # Транспонируем матрицу
        transposed_matrix = list(map(list, zip(*matrix)))
        
        # Создаем новую структуру данных
        transposed_data = []
        processed_cells = set()
        
        for i, row in enumerate(transposed_matrix):
            new_row = []
            for j, cell in enumerate(row):
                if cell is not None and (i, j) not in processed_cells:
                    # Создаем новую ячейку с поменянными местами rowspan и colspan
                    new_cell = {
                        'value': cell['value'],
                        'rowspan': cell['colspan'],  # Меняем местами
                        'colspan': cell['rowspan'],  # Меняем местами
                        'nottitle': cell['nottitle']
                    }
                    new_row.append(new_cell)
                    
                    # Отмечаем обработанные ячейки
                    for r in range(new_cell['rowspan']):
                        for c in range(new_cell['colspan']):
                            processed_cells.add((i + r, j + c))
            
            if new_row:  # Добавляем строку только если в ней есть ячейки
                transposed_data.append(new_row)

        table_span = SpanTable()
        table_span.set_table(transposed_data)
        return table_span

    def get_tag_structure(self, element):
        """Получает структуру тегов элемента в виде списка"""

        structure = []

        if isinstance(element, str):
            return None
        if element.name == 'span':
            return ''
        structure.append(element.name)

        for child in element.children:
            if isinstance(child, str) and child.strip() == '':
                continue
            
            child_structure = self.get_tag_structure(child)

            if child_structure:
                structure.append(child_structure)

        return structure

    def vertical_check(self):
        """Проверка на подлиность вертикальных таблиц"""
        
        table_spans = self.get_copy().get_table()
        if len(table_spans)<2:
            return False
        
        old = []
        new = []
        
        for cell in table_spans[0]:
            old += [cell]
                
        for i in range(1,len(table_spans)):
            j = 0
            k = 0
            s = 0
            
            while (k < len(old)):
                
                if (old[k]["rowspan"] > 1):
                    old[k]["rowspan"] -= 1
                    new += [old[k]]
                    k += 1
                    continue
                
                if (j < len(table_spans[i])):
                    
                    if (s + table_spans[i][j]["colspan"]) < old[k]["colspan"]:
                        s += table_spans[i][j]["colspan"]
                        new += [table_spans[i][j]]
                        j += 1
                        continue
                    
                    if (s + table_spans[i][j]["colspan"]) == old[k]["colspan"]:
                        if s == 0:
                            table_spans[i][j]["nottitle"] = True
                            if old[k]["nottitle"] == True:
                                old_structure = self.get_tag_structure(old[k]['value'])
                                new_structure = self.get_tag_structure(table_spans[i][j]['value'])
                                if not(old_structure == new_structure):
                                    return False
                        s = 0
                        new += [table_spans[i][j]]
                        j += 1
                        k += 1
                        continue
                    
                return False

            if not((j == len(table_spans[i])) and (k == len(old))):
                return False
            old = [cell for cell in new]
            new =[]

        if not(all(cell['rowspan'] == 1 for cell in old)):
            return False
        return True

    def is_top(self):
        return self.vertical_check()

    def is_left(self):
        table_span = self.get_transpose()
        return table_span.vertical_check()

    def is_right(self):
        table_span = self.get_transpose().get_flip()
        return table_span.vertical_check()

    def is_bottom(self):
        table_span = self.get_transpose().get_flip().get_transpose()
        return table_span.vertical_check()
        


def write_to_excel(output_excel_path, tables):
    """Запись таблицы в excel-фил"""
    wb = Workbook()

    # Удаляем дефолтную пустую страницу, чтобы потом создавать по одной на каждый table
    while len(wb.worksheets) > 0:
        wb.remove(wb.worksheets[0])

    for idx, table in enumerate(tables, start=1):
        sheet_name = f"Таблица_{idx}"
        ws = wb.create_sheet(title=sheet_name)

        # Получаем все строки данной таблицы
        rows = table.find_all("tr")
        # Матрица для хранения инфо о занятых ячейках
        occupied = {}

        current_row = 1
        for row in rows:
            if row.decode_contents() == '':
                continue
            # Получаем все столбцы (теги td и th) в строке
            cells = row.find_all(["td", "th"])
            current_col = 1
            for cell in cells:
                # Пропускаем уже занятые ячейки
                while (current_row, current_col) in occupied:
                    current_col += 1

                rowspan = int(cell.get("rowspan", 1))
                colspan = int(cell.get("colspan", 1))

                # Мёржим ячейки, если указан rowspan/colspan
                if rowspan > 1 or colspan > 1:
                    start_row = current_row
                    start_col = current_col
                    end_row = current_row + rowspan - 1
                    end_col = current_col + colspan - 1
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=start_col,
                        end_row=end_row,
                        end_column=end_col
                    )

                    # Помечаем все объединённые ячейки как занятые
                    for r in range(start_row, end_row + 1):
                        for c in range(start_col, end_col + 1):
                            occupied[(r, c)] = True

                value = cell.get_text(strip=True)
                ws.cell(row=current_row, column=current_col, value=value)

                current_col += 1

            current_row += 1

    wb.save(output_excel_path)
    
def get_tables(html_path, format_table):
    """Извлечение таблицы"""
    if format_table == "file":

        if not os.path.isfile(html_path):
            print(f"Файл {html_path} не найден.")
            sys.exit(1)
        
        with open(html_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f, "html.parser")
        
    elif format_table == "url":
        f = download_html(html_path)
        soup = BeautifulSoup(f, "html.parser")
    else:
        print("воходные данные не коректны")
        sys.exit(1)
    tables = soup.find_all("table")
    return tables

def download_html(url):
    """Скачивание HTML содержимого по-заданному URL."""
    response = requests.get(url, verify=False)
    response.raise_for_status()  # Проверка успешности запроса
    return response.text

def check(table):
    """Проверка на подлинность"""
    table_span = SpanTable()  # Исправлено: правильная инициализация
    table_span.make_table(table)
    
    if table_span.is_top():
        print("top")
        return True
    
    if table_span.is_left():
        print("left")
        return True
    
    if table_span.is_right():
        print("right")
        return True
    
    if table_span.is_bottom():
        print("bottom")
        return True
    
    print("not")
    return False

def get_genuine_tables(tables):
    """Получение все подленных таблиц"""
    genuine_tables=[]
    for table in tables:
        if not(table.find("table")) and check(table):
            genuine_tables+=[table]
    return genuine_tables

def data_acquisition():
    """Получение входных данных"""
    if len(sys.argv) < -1:#3:
        print("Использование: python script.py url/file путь_к_html путь_к_xlsx")
        sys.exit(1)
    format_table = 'file' #sys.argv[1]
    html_path = 'table.html'#sys.argv[2]
    xlsx_path = 'example.xlsx'#sys.argv[3]

    tables = get_tables(html_path, format_table)

    return {'html_path': html_path, 'xlsx_path': xlsx_path, 'tables': tables}

if __name__ == "__main__":
    data = data_acquisition()
    genuine_tables = get_genuine_tables(data['tables'])
    write_to_excel(data['xlsx_path'], genuine_tables)
    os.startfile(data['xlsx_path'])

    

    
